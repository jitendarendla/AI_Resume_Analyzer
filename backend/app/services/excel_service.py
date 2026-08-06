import os
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.config import settings
from app.services.parser_service import (
    clean_candidate_name,
    clean_candidate_location,
    extract_technology_title,
    extract_rule_based_details,
    extract_all_skills_comprehensively
)

def generate_excel_report(report_name: str, candidate_data: list[dict]) -> str:
    filename = f"{report_name.replace(' ', '_')}_{int(datetime.now().timestamp())}.xlsx"
    filepath = os.path.join(settings.REPORTS_DIR, filename)

    rows = []
    for idx, item in enumerate(candidate_data, start=1):
        raw_name = item.get("name") or "Candidate"
        email = item.get("email") or ""
        phone = item.get("phone") or ""
        resume_file = item.get("file_name") or item.get("resume_file") or ""
        raw_loc = item.get("location") or ""
        raw_text = item.get("raw_text") or ""
        skills_raw = item.get("skills") or []

        comp_skills = extract_all_skills_comprehensively(raw_text) if raw_text else []

        if isinstance(skills_raw, list):
            all_skills = list(dict.fromkeys(skills_raw + comp_skills))
        elif isinstance(skills_raw, str) and skills_raw.strip():
            all_skills = list(dict.fromkeys([s.strip() for s in skills_raw.split(",") if s.strip()] + comp_skills))
        else:
            all_skills = comp_skills

        skills_str = ", ".join([str(s) for s in all_skills if s])

        clean_name = clean_candidate_name(raw_name, resume_file, email, raw_text)
        clean_loc = clean_candidate_location(raw_loc, raw_text)
        tech_title = extract_technology_title(raw_text, resume_file, all_skills)

        rows.append({
            "S.No": idx,
            "Candidate Name": clean_name,
            "Email ID": email,
            "Phone Number": phone,
            "Location": clean_loc,
            "Technology/Title": tech_title,
            "Skills": skills_str
        })

    if not rows:
        rows.append({
            "S.No": 1,
            "Candidate Name": "N/A",
            "Email ID": "N/A",
            "Phone Number": "N/A",
            "Location": "N/A",
            "Technology/Title": "N/A",
            "Skills": "N/A"
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidate Summary"

    # Show gridlines
    ws.views.sheetView[0].showGridLines = True

    # Freeze header row
    ws.freeze_panes = "A2"

    # Premium Style System (Executive Navy Theme)
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    odd_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    name_font = Font(name="Segoe UI", size=10.5, bold=True, color="0F172A")
    title_font = Font(name="Segoe UI", size=10, bold=True, color="0369A1")
    email_font = Font(name="Segoe UI", size=10, color="2563EB", underline="single")
    data_font = Font(name="Segoe UI", size=10, color="1E293B")
    skills_font = Font(name="Segoe UI", size=9.5, color="334155")

    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    header_border = Border(
        left=Side(style="thin", color="1E293B"),
        right=Side(style="thin", color="1E293B"),
        top=Side(style="medium", color="0F172A"),
        bottom=Side(style="medium", color="0284C7")
    )

    headers = ["S.No", "Candidate Name", "Email ID", "Phone Number", "Location", "Technology/Title", "Skills"]
    ws.append(headers)
    ws.row_dimensions[1].height = 34

    for col_num, header_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(rows, start=2):
        ws.append([
            row_data["S.No"],
            row_data["Candidate Name"],
            row_data["Email ID"],
            row_data["Phone Number"],
            row_data["Location"],
            row_data["Technology/Title"],
            row_data["Skills"]
        ])
        ws.row_dimensions[row_idx].height = 28
        fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill

        for col_num in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.fill = fill
            cell.border = thin_border
            
            if col_num == 1:  # S.No
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = data_font
            elif col_num == 2:  # Candidate Name
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = name_font
            elif col_num == 3:  # Email ID
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = email_font if "@" in str(cell.value) else data_font
            elif col_num in [4, 5]:  # Phone, Location
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = data_font
            elif col_num == 6:  # Technology/Title
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.font = title_font
            elif col_num == 7:  # Skills
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.font = skills_font

    # Set explicit column widths for beautiful layout
    col_widths = {
        "A": 9,   # S.No
        "B": 28,  # Candidate Name
        "C": 32,  # Email ID
        "D": 18,  # Phone Number
        "E": 24,  # Location
        "F": 34,  # Technology/Title
        "G": 65   # Skills
    }

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    os.makedirs(settings.REPORTS_DIR, exist_ok=True)
    wb.save(filepath)
    return filepath
